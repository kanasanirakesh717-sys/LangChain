import os
import docx
import openpyxl
import faiss
import numpy as np
from PyPDF2 import PdfReader
from dotenv import load_dotenv

# LangChain imports (Gemini)
from langchain_google_genai import GoogleGenerativeAIEmbeddings, ChatGoogleGenerativeAI

# ---------------- Config ----------------
CHUNK_SIZE_WORDS = 250
TOP_K_CHUNKS = 3

# Load environment variables from .env (if present) and read API key
load_dotenv()
API_KEY = os.getenv("GOOGLE_API_KEY")
if not API_KEY:
    raise RuntimeError(
        "Missing API key. Please set GOOGLE_API_KEY or OPENAI_API_KEY in your environment or in a .env file."
    )

# Ensure the expected env vars are set for downstream libraries
os.environ.setdefault("GOOGLE_API_KEY", API_KEY)

# ---------------- Load Gemini Models ----------------

print("[INFO] Loading Gemini embeddings and LLM...")

embed_model = GoogleGenerativeAIEmbeddings(model="models/embedding-001", transport="rest")
print(embed_model.embed_query("Hello world!"))
llm = ChatGoogleGenerativeAI(model="gemini-2.5-flash",transport='rest')

# -------- File Text Extraction --------
def extract_text_from_pdf(file_obj):
    reader = PdfReader(file_obj)
    return "\n".join(page.extract_text() for page in reader.pages if page.extract_text())

def extract_text_from_docx(file_obj):
    document = docx.Document(file_obj)
    return "\n".join(p.text for p in document.paragraphs if p.text.strip())

def extract_text_from_xlsx(file_obj):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    return "\n".join(
        " ".join(str(cell) for cell in row if cell is not None)
        for sheet in wb.sheetnames
        for row in wb[sheet].iter_rows(values_only=True)
    )

def extract_text(file_path):
    name = file_path.lower()
    with open(file_path, "rb") as f:
        if name.endswith(".pdf"):
            return extract_text_from_pdf(f)
        elif name.endswith(".docx"):
            return extract_text_from_docx(f)
        elif name.endswith(".xlsx"):
            return extract_text_from_xlsx(f)
        else:
            return ""

# ---------------- Text Chunking ----------------
def chunk_texts(text, source, chunk_size_words=CHUNK_SIZE_WORDS):
    words = text.split()
    return [
        {"text": " ".join(words[i:i + chunk_size_words]), "meta": {"source_file": source}}
        for i in range(0, len(words), chunk_size_words)
    ]

# ---------------- FAISS Index ----------------
def build_faiss_index(chunks):
    texts = [c["text"] for c in chunks]
    embeddings = embed_model.embed_documents(texts)
    embeddings = np.array(embeddings, dtype="float32")
    dim = embeddings.shape[1]

    index = faiss.IndexFlatL2(dim)
    index.add(embeddings)

    return index, embeddings

def search_chunks(index, query, chunks, top_k=TOP_K_CHUNKS):
    query_emb = embed_model.embed_query(query)
    query_emb = np.array([query_emb], dtype="float32")

    distances, indices = index.search(query_emb, top_k)

    results = []
    for idx, dist in zip(indices[0], distances[0]):
        results.append({
            "chunk": chunks[idx],
            "score": float(dist)
        })
    return results

# -------- Main Terminal App --------
if __name__ == "__main__":
    print("📚 Simple RAG with Gemini + Manual FAISS")

    file_paths = input("Enter file paths separated by commas: ").split(",")
    file_paths = [fp.strip() for fp in file_paths if fp.strip()]

    if not file_paths:
        print("[ERROR] No files provided.")
        exit()

    query = input("Enter your question: ").strip()
    if not query:
        print("[ERROR] No question provided.")
        exit()

    print("[INFO] Extracting and chunking text...")
    chunks = []
    for path in file_paths:
        if not os.path.exists(path):
            print(f"[WARNING] File not found: {path}")
            continue
        text = extract_text(path)
        if text.strip():
            chunks.extend(chunk_texts(text, os.path.basename(path)))

    if not chunks:
        print("[ERROR] No text extracted from files.")
        exit()

    print("[INFO] Building FAISS index...")
    index, _ = build_faiss_index(chunks)

    print("[INFO] Searching for relevant chunks...")
    retrieved = search_chunks(index, query, chunks)

    print("[INFO] Running Gemini QA...")
    context = "\n\n".join([r["chunk"]["text"] for r in retrieved])
    prompt = (
        f"Answer the following question in a complete and detailed sentence "
        f"based only on the provided context.\n\n"
        f"Context:\n{context}\n\nQuestion: {query}\n\n"
    )

    result = llm.invoke(prompt)
    answer = result.content.strip()

    print("\n🔎 Answer:")
    if answer:
        print(f"✅ {answer}")
    else:
        print("ℹ️ No exact answer found.")

    print(f"\n📜 Top {TOP_K_CHUNKS} Relevant Chunks:")
    for r in retrieved:
        print(f"\n--- File: {r['chunk']['meta']['source_file']} (score={r['score']:.3f}) ---")
        print(r['chunk']['text'])
